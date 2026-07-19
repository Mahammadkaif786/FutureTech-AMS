from flask import Blueprint, render_template, redirect, url_for, flash, request, make_response, g
from flask_login import login_required
from datetime import datetime, date, timedelta
from models import Student, Course, Payment, db
import io
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)

def get_report_data(report_type, start_date=None, end_date=None, course_id=None):
    """
    Core engine to filter and extract report datasets.
    """
    students_query = Student.query
    payments_query = Payment.query.join(Student)

    # 1. Base Filters
    if course_id:
        students_query = students_query.filter(Student.course_id == course_id)
        payments_query = payments_query.filter(Student.course_id == course_id)
        
    if start_date:
        students_query = students_query.filter(Student.admission_date >= start_date)
        payments_query = payments_query.filter(Payment.payment_date >= datetime.combine(start_date, datetime.min.time()))
        
    if end_date:
        students_query = students_query.filter(Student.admission_date <= end_date)
        payments_query = payments_query.filter(Payment.payment_date <= datetime.combine(end_date, datetime.max.time()))

    # 2. Segmenting reports
    if report_type == 'all_students':
        data = students_query.order_by(Student.student_id.desc()).all()
        headers = ['Student ID', 'Full Name', 'Mobile', 'Course', 'Admission Date', 'Total Paid', 'Pending Dues', 'Status']
        
    elif report_type == 'active_students':
        data = students_query.filter(Student.status == 'Active').order_by(Student.student_id.desc()).all()
        headers = ['Student ID', 'Full Name', 'Mobile', 'Course', 'Admission Date', 'Total Paid', 'Pending Dues']
        
    elif report_type == 'completed_students':
        data = students_query.filter(Student.status == 'Completed').order_by(Student.completion_date.desc()).all()
        headers = ['Student ID', 'Full Name', 'Course', 'Admission Date', 'Completion Date', 'Certificate ID']
        
    elif report_type == 'course_wise':
        data = students_query.order_by(Student.course_id, Student.student_id.desc()).all()
        headers = ['Course', 'Student ID', 'Full Name', 'Mobile', 'Admission Date', 'Total Fee', 'Total Paid', 'Status']
        
    elif report_type == 'fee_collection':
        data = payments_query.order_by(Payment.payment_date.desc()).all()
        headers = ['Receipt ID', 'Student ID', 'Student Name', 'Course', 'Payment Date', 'Payment Method', 'Amount Collected']
        
    elif report_type == 'pending_fees':
        data = students_query.filter(Student.pending_fee > 0).order_by(Student.pending_fee.desc()).all()
        headers = ['Student ID', 'Full Name', 'Mobile', 'Course', 'Total Fee', 'Total Paid', 'Pending Dues']
        
    elif report_type == 'monthly_admissions':
        # Summary report: Count admissions per month
        data = students_query.order_by(Student.admission_date.desc()).all()
        headers = ['Month/Year', 'Admissions Count', 'Course Revenue Potential']
    else:
        data = []
        headers = []

    return data, headers

@reports_bp.route('/reports')
@login_required
def index():
    report_type = request.args.get('report_type', 'all_students')
    course_id = request.args.get('course_id', type=int)
    
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    start_date = None
    end_date = None
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
            
    courses = Course.query.order_by(Course.course_name).all()
    
    # Fetch report datasets
    report_data, headers = get_report_data(report_type, start_date, end_date, course_id)
    
    # Summary items for Monthly Admissions grouping in Python
    monthly_summary = {}
    if report_type == 'monthly_admissions' and report_data:
        for s in report_data:
            month_label = f"{calendar.month_name[s.admission_date.month]} {s.admission_date.year}"
            if month_label not in monthly_summary:
                monthly_summary[month_label] = {'count': 0, 'revenue': 0.0}
            monthly_summary[month_label]['count'] += 1
            monthly_summary[month_label]['revenue'] += s.total_fee
            
    return render_template('reports/list.html',
                           courses=courses,
                           report_type=report_type,
                           course_id=course_id,
                           start_date_str=start_date_str,
                           end_date_str=end_date_str,
                           report_data=report_data,
                           headers=headers,
                           monthly_summary=monthly_summary,
                           active_page='reports')

@reports_bp.route('/reports/export/excel')
@login_required
def export_excel():
    report_type = request.args.get('report_type', 'all_students')
    course_id = request.args.get('course_id', type=int)
    
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    start_date = None
    end_date = None
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
            
    report_data, headers = get_report_data(report_type, start_date, end_date, course_id)
    
    # 1. Create openpyxl workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Registry Report"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styling variables
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # Deep Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True, color="1E40AF")
    subtitle_font = Font(name="Calibri", size=11, italic=True)
    
    # 2. Add Title Blocks
    ws.append([])
    ws.cell(row=2, column=2, value=f"{g.settings.institute_name if g.settings else 'FutureTech Computer Institute'}").font = title_font
    
    report_titles = {
        'all_students': 'All Registered Students Report',
        'active_students': 'Active Enrolled Students Report',
        'completed_students': 'Course Graduates Report',
        'course_wise': 'Course-wise Enrollment Report',
        'fee_collection': 'Fee Payment Collections Report',
        'pending_fees': 'Pending Fees Outstanding Report',
        'monthly_admissions': 'Monthly Admissions & Revenue Summary'
    }
    
    ws.cell(row=3, column=2, value=f"{report_titles.get(report_type, 'Student Registry Report')}").font = Font(name="Calibri", size=13, bold=True)
    
    range_str = f"Date Range: {start_date_str or 'Beginning'} to {end_date_str or 'Present'}"
    ws.cell(row=4, column=2, value=range_str).font = subtitle_font
    
    ws.append([]) # spacer
    ws.append([]) # spacer
    
    # 3. Add Headers
    ws.append(headers)
    header_row_num = 7
    
    # Apply header formatting
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_num, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Thin borders helper
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # 4. Populate Data Rows
    if report_type == 'monthly_admissions':
        # Group monthly summary
        monthly_summary = {}
        for s in report_data:
            month_label = f"{calendar.month_name[s.admission_date.month]} {s.admission_date.year}"
            if month_label not in monthly_summary:
                monthly_summary[month_label] = {'count': 0, 'revenue': 0.0}
            monthly_summary[month_label]['count'] += 1
            monthly_summary[month_label]['revenue'] += s.total_fee
            
        row_num = header_row_num + 1
        for month_label, stats in monthly_summary.items():
            row_data = [month_label, stats['count'], stats['revenue']]
            ws.append(row_data)
            
            # format numeric revenue cells
            ws.cell(row=row_num, column=3).number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center")
            row_num += 1
            
    else:
        row_num = header_row_num + 1
        for item in report_data:
            row_data = []
            
            if report_type == 'all_students':
                row_data = [
                    item.student_id, item.full_name, item.mobile, 
                    item.course.course_name, item.admission_date.strftime('%Y-%m-%d'),
                    item.paid_fee, item.pending_fee, item.status
                ]
            elif report_type == 'active_students':
                row_data = [
                    item.student_id, item.full_name, item.mobile, 
                    item.course.course_name, item.admission_date.strftime('%Y-%m-%d'),
                    item.paid_fee, item.pending_fee
                ]
            elif report_type == 'completed_students':
                row_data = [
                    item.student_id, item.full_name, item.course.course_name,
                    item.admission_date.strftime('%Y-%m-%d'), 
                    item.completion_date.strftime('%Y-%m-%d') if item.completion_date else '',
                    item.certificate_number or ''
                ]
            elif report_type == 'course_wise':
                row_data = [
                    item.course.course_name, item.student_id, item.full_name, item.mobile,
                    item.admission_date.strftime('%Y-%m-%d'), item.total_fee, item.paid_fee, item.status
                ]
            elif report_type == 'fee_collection':
                row_data = [
                    item.receipt_number, item.student.student_id, item.student.full_name,
                    item.student.course.course_name, item.payment_date.strftime('%Y-%m-%d %H:%M'),
                    item.payment_method, item.amount
                ]
            elif report_type == 'pending_fees':
                row_data = [
                    item.student_id, item.full_name, item.mobile, item.course.course_name,
                    item.total_fee, item.paid_fee, item.pending_fee
                ]
                
            ws.append(row_data)
            
            # Formats & alignments per report type
            if report_type == 'all_students':
                ws.cell(row=row_num, column=6).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=7).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
            elif report_type == 'active_students':
                ws.cell(row=row_num, column=6).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=7).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
            elif report_type == 'completed_students':
                ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
            elif report_type == 'course_wise':
                ws.cell(row=row_num, column=6).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=7).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
            elif report_type == 'fee_collection':
                ws.cell(row=row_num, column=7).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
            elif report_type == 'pending_fees':
                ws.cell(row=row_num, column=5).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=6).number_format = '₹#,##0.00'
                ws.cell(row=row_num, column=7).number_format = '₹#,##0.00'

            row_num += 1
            
    # Apply standard fonts and borders to all populated rows
    for r in range(header_row_num, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            if r > header_row_num:
                cell.font = data_font
            cell.border = thin_border
            
    # 5. Auto-fit Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't size based on title rows 2-4
        for cell in col:
            if cell.row >= header_row_num and cell.value:
                max_len = max(max_len, len(str(cell.value)))
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save workbook to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"FTCI_Report_{report_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response
